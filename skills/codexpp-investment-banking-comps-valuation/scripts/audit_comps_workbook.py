#!/usr/bin/env python3
"""Audit a comparable company analysis workbook for structural and formula issues.

Usage:
    python audit_comps_workbook.py model.xlsx --json-out audit.json --markdown-out audit.md
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

EXPECTED_SHEET_GROUPS = {
    "control": ["control", "input", "assumption"],
    "universe": ["universe", "peer", "comp"],
    "market_data": ["market", "price", "ev", "capitalization"],
    "financials": ["financial", "statement", "estimate", "consensus"],
    "adjustments": ["adjust", "normal", "calendar"],
    "multiples": ["multiple", "valuation metric"],
    "benchmarking": ["benchmark", "kpi", "operating"],
    "valuation": ["valuation", "output", "summary"],
    "sources": ["source", "citation", "footnote"],
    "qa": ["qa", "check", "audit"],
}

ERROR_PATTERNS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#N/A")
SOURCE_TERMS = (
    "source",
    "retrieval",
    "as of",
    "data date",
    "valuation date",
    "filing",
    "accession",
    "url",
)
REQUIREMENTS_FILE = Path(__file__).resolve().parent / "requirements.txt"


def require_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "Missing optional workbook dependency: openpyxl.\n"
            f"Install local script dependencies with: python3 -m pip install -r {REQUIREMENTS_FILE}\n"
            "Manual fallback: inspect workbook structure, formulas, source logs, "
            "external links, hardcodes, peer duplicates, and QA evidence using "
            "references/qa-and-pressure-testing.md, then state that automated audit "
            "was not run."
        )
    return load_workbook


def normalize(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", name.lower()).strip()


def sheet_matches(sheet_name: str, terms: list[str]) -> bool:
    normalized = normalize(sheet_name)
    return any(term in normalized for term in terms)


def find_sheet_groups(sheet_names: list[str]) -> dict[str, list[str]]:
    matches: dict[str, list[str]] = {}
    for group, terms in EXPECTED_SHEET_GROUPS.items():
        matches[group] = [name for name in sheet_names if sheet_matches(name, terms)]
    return matches


def used_cells(ws) -> list[Any]:
    return [cell for row in ws.iter_rows() for cell in row if cell.value not in (None, "")]


def audit_workbook(path: Path) -> dict[str, Any]:
    load_workbook = require_openpyxl()
    wb = load_workbook(path, data_only=False, read_only=False)
    sheet_names = wb.sheetnames
    groups = find_sheet_groups(sheet_names)

    findings: list[dict[str, str]] = []
    metrics: dict[str, Any] = {
        "workbook": str(path),
        "audited_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "sheet_count": len(sheet_names),
        "sheets": sheet_names,
        "matched_sheet_groups": groups,
    }

    missing_groups = [group for group, names in groups.items() if not names]
    if missing_groups:
        findings.append(
            {
                "severity": "warning",
                "category": "structure",
                "finding": "Missing or unclear expected sheet groups: " + ", ".join(missing_groups),
                "recommendation": "Add or clearly label tabs for these functions, or document where they are handled.",
            }
        )

    formula_counts = Counter()
    hardcode_counts = Counter()
    error_cells: list[str] = []
    external_formula_cells: list[str] = []
    formula_by_sheet: dict[str, int] = {}
    numeric_by_sheet = Counter()
    comments_by_sheet: dict[str, int] = {}
    source_term_hits: list[str] = []

    ticker_values: list[str] = []
    ticker_headers_seen = False

    for ws in wb.worksheets:
        cells = used_cells(ws)
        formula_count = 0
        numeric_count = 0
        comments_count = 0
        for cell in cells:
            value = cell.value
            coord = f"{ws.title}!{cell.coordinate}"
            if cell.comment:
                comments_count += 1
            if isinstance(value, str):
                lower = value.lower()
                if any(term in lower for term in SOURCE_TERMS):
                    source_term_hits.append(coord)
                if value.startswith("="):
                    formula_count += 1
                    formula_counts[ws.title] += 1
                    if any(pattern in value.upper() for pattern in ERROR_PATTERNS):
                        error_cells.append(coord)
                    if "[" in value and "]" in value:
                        external_formula_cells.append(coord)
                elif any(pattern in value.upper() for pattern in ERROR_PATTERNS):
                    error_cells.append(coord)
            elif isinstance(value, (int, float)):
                numeric_count += 1
                numeric_by_sheet[ws.title] += 1
                if (
                    any(
                        key in normalize(ws.title)
                        for key in ("valuation", "output", "summary", "multiple")
                    )
                    and cell.row > 3
                ):
                    hardcode_counts[ws.title] += 1
            if isinstance(value, str) and normalize(value) in ("ticker", "tickers"):
                ticker_headers_seen = True
                # collect values under this header in the same column
                for r in range(cell.row + 1, min(ws.max_row, cell.row + 60) + 1):
                    maybe = ws.cell(r, cell.column).value
                    if isinstance(maybe, str) and maybe.strip() and not maybe.startswith("="):
                        ticker_values.append(maybe.strip().upper())
        formula_by_sheet[ws.title] = formula_count
        numeric_by_sheet[ws.title] = numeric_count
        comments_by_sheet[ws.title] = comments_count

        if any(key in normalize(ws.title) for key in ("valuation", "output", "summary")):
            if formula_count == 0 and numeric_count > 5:
                findings.append(
                    {
                        "severity": "warning",
                        "category": "formula_integrity",
                        "finding": f"{ws.title} has numeric values but no formulas detected.",
                        "recommendation": "Review whether valuation outputs are hardcoded instead of linked to assumptions and data tabs.",
                    }
                )

    metrics["formula_counts_by_sheet"] = formula_by_sheet
    metrics["numeric_counts_by_sheet"] = dict(numeric_by_sheet)
    metrics["comments_by_sheet"] = comments_by_sheet
    metrics["external_formula_cells"] = external_formula_cells[:100]
    metrics["error_cells"] = error_cells[:100]
    metrics["hardcode_counts_on_output_like_sheets"] = dict(hardcode_counts)

    if error_cells:
        findings.append(
            {
                "severity": "critical",
                "category": "formula_integrity",
                "finding": f"Detected formula or cell error patterns in {len(error_cells)} cells.",
                "recommendation": "Inspect and fix error cells before senior review. Do not rely on affected outputs.",
            }
        )

    if external_formula_cells:
        findings.append(
            {
                "severity": "warning",
                "category": "formula_integrity",
                "finding": f"Detected external workbook references in {len(external_formula_cells)} formula cells.",
                "recommendation": "Break, document, or preserve external links intentionally. Final deliverables should not depend on hidden external files unless requested.",
            }
        )

    if hardcode_counts:
        findings.append(
            {
                "severity": "warning",
                "category": "formula_integrity",
                "finding": "Output-like sheets contain numeric hardcodes: "
                + ", ".join(f"{k}={v}" for k, v in hardcode_counts.items()),
                "recommendation": "Confirm these are labeled assumptions. Replace hidden hardcoded outputs with formulas where possible.",
            }
        )

    if not source_term_hits:
        findings.append(
            {
                "severity": "warning",
                "category": "sources",
                "finding": "No obvious source/date terminology detected in workbook cells.",
                "recommendation": "Add a Sources tab with source name, retrieval date, data date, metric, period, and confidence label.",
            }
        )

    if ticker_headers_seen:
        dupes = [ticker for ticker, count in Counter(ticker_values).items() if count > 1]
        if dupes:
            findings.append(
                {
                    "severity": "warning",
                    "category": "peer_universe",
                    "finding": "Duplicate ticker values detected: " + ", ".join(dupes[:20]),
                    "recommendation": "Check for duplicate peers, repeated target rows, or accidental copy/paste errors.",
                }
            )
    else:
        findings.append(
            {
                "severity": "info",
                "category": "peer_universe",
                "finding": "No explicit ticker header detected.",
                "recommendation": "Ensure the peer universe has a clear ticker/company identifier column.",
            }
        )

    # Check whether source and QA tabs have meaningful comments/logs.
    comments_total = sum(comments_by_sheet.values())
    if comments_total == 0:
        findings.append(
            {
                "severity": "info",
                "category": "documentation",
                "finding": "No cell comments detected.",
                "recommendation": "Consider adding comments to major assumptions and sourced inputs for auditability.",
            }
        )

    # Scoring: start at 100 and subtract by severity.
    score = 100
    penalty = {"critical": 25, "warning": 10, "info": 3}
    for item in findings:
        score -= penalty.get(item["severity"], 0)
    metrics["quality_score_directional"] = max(score, 0)

    return {"metrics": metrics, "findings": findings}


def to_markdown(report: dict[str, Any]) -> str:
    metrics = report["metrics"]
    findings = report["findings"]
    by_severity = defaultdict(list)
    for item in findings:
        by_severity[item["severity"]].append(item)

    lines = [
        "# Comps Workbook Audit",
        "",
        f"Workbook: `{metrics['workbook']}`",
        f"Audited at: {metrics['audited_at']}",
        f"Directional quality score: {metrics['quality_score_directional']}/100",
        "",
        "## Sheet coverage",
        "",
    ]
    for group, sheets in metrics["matched_sheet_groups"].items():
        status = ", ".join(sheets) if sheets else "missing/unclear"
        lines.append(f"- {group}: {status}")
    lines.extend(["", "## Findings", ""])
    if not findings:
        lines.append(
            "No issues detected by automated checks. Manual MD/PM review is still required."
        )
    else:
        for severity in ("critical", "warning", "info"):
            if severity in by_severity:
                lines.append(f"### {severity.title()}")
                lines.append("")
                for item in by_severity[severity]:
                    lines.append(
                        f"- **{item['category']}**: {item['finding']} Recommendation: {item['recommendation']}"
                    )
                lines.append("")
    lines.extend(
        [
            "## Notes",
            "",
            "This automated audit checks structure, formulas, external links, obvious hardcodes, source terminology, comments, and duplicate tickers. It does not replace manual peer selection, normalization, or valuation judgment review.",
        ]
    )
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Audit a comps workbook for structural and formula issues."
    )
    parser.add_argument("workbook", help="Path to .xlsx workbook.")
    parser.add_argument("--json-out", help="Optional path for JSON audit output.")
    parser.add_argument("--markdown-out", help="Optional path for Markdown audit output.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    path = Path(args.workbook).resolve()
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")
    report = audit_workbook(path)
    if args.json_out:
        Path(args.json_out).write_text(json.dumps(report, indent=2), encoding="utf-8")
    if args.markdown_out:
        Path(args.markdown_out).write_text(to_markdown(report), encoding="utf-8")
    print(to_markdown(report))


if __name__ == "__main__":
    main()
