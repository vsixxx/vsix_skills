#!/usr/bin/env python3
"""Conservative first-pass cleaner for messy CSV/XLSX data.

Creates an audit-ready workbook with cleaned data, raw source, data dictionary,
quality checks, and assumptions/audit sheets. This is a helper for the skill's
analyst workflow, not a replacement for human/AI review of ambiguous business rules.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

PLUGIN_ROOT = Path(__file__).resolve().parents[5]
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from shared.artifacts import write_artifact_manifest  # noqa: E402

BLANK_TOKENS = {"", "-", "--", "n/a", "na", "null", "none", "not available", "#n/a", "nan"}
TOTAL_RE = re.compile(r"\b(grand\s+total|subtotal|sub-total|total)\b", re.I)
ID_HINT_RE = re.compile(
    r"\b(id|code|sku|zip|postal|phone|cusip|isin|sedol|ticker|account\s*#|invoice\s*#|po\s*#|employee)\b",
    re.I,
)
CURRENCY_RE = re.compile(r"[$€£¥]|\b(usd|eur|gbp|jpy|cad|aud|chf|cny)\b", re.I)
PERCENT_RE = re.compile(r"%|\bbps\b|\bbasis\s+points\b", re.I)
NUMERIC_RE = re.compile(
    r"^\(?[+-]?\s*[$€£¥]?\s*\d{1,3}(?:,\d{3})*(?:\.\d+)?\s*\)?$|^\(?[+-]?\s*[$€£¥]?\s*\d+(?:\.\d+)?\s*\)?$"
)

DOMAIN_REQUIRED_HINTS: dict[str, list[str]] = {
    "finance": ["amount", "account", "period", "scenario", "department", "currency"],
    "capiq_factset_export": ["ticker", "company", "period", "revenue", "ebitda", "currency"],
    "cim_table": ["metric", "period", "revenue", "ebitda", "adjustment", "source"],
    "debt_schedule": ["facility", "maturity", "rate", "spread", "debt", "amortization"],
    "process_tracker": ["buyer", "investor", "status", "owner", "nda", "ioi", "loi"],
    "qoe_schedule": ["adjustment", "ebitda", "quality", "source", "period", "basis"],
    "investing": ["ticker", "security", "date", "price", "return", "currency"],
    "operations": ["id", "owner", "status", "priority", "created", "resolved"],
    "sales": ["account", "opportunity", "stage", "owner", "close", "amount"],
    "product": ["user", "event", "timestamp", "session", "platform"],
    "hr": ["employee", "manager", "status", "start", "location"],
    "procurement": ["vendor", "invoice", "po", "contract", "renewal", "amount", "currency"],
}

OPENPYXL_INSTALL_COMMAND = "python3 -m pip install -r scripts/requirements.txt"
OPENPYXL_REVIEW_FALLBACK = (
    "Manual-review fallback: export workbook sheets to CSV or inspect them in Excel/Sheets, "
    "then preserve the raw source, data dictionary, quality checks, and assumptions/audit notes while cleaning manually."
)
_OPENPYXL_MODULES: tuple[Any, Any, Any, Any] | None = None


class DependencyPreflightError(RuntimeError):
    """Raised when an optional workbook dependency is missing."""


def openpyxl_dependency_message(action: str) -> str:
    return (
        f"openpyxl is required to {action}.\n"
        f"Install locally from this skill directory with: {OPENPYXL_INSTALL_COMMAND}\n"
        f"{OPENPYXL_REVIEW_FALLBACK}"
    )


def load_openpyxl_modules() -> tuple[Any, Any, Any, Any]:
    global _OPENPYXL_MODULES
    if _OPENPYXL_MODULES is not None:
        return _OPENPYXL_MODULES
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ModuleNotFoundError as exc:
        if exc.name == "openpyxl":
            raise DependencyPreflightError(
                openpyxl_dependency_message("write and format the cleaned workbook")
            ) from exc
        raise
    _OPENPYXL_MODULES = (load_workbook, Table, TableStyleInfo, get_column_letter)
    return _OPENPYXL_MODULES


def norm_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ").strip())


def is_blank(value: Any) -> bool:
    return norm_cell(value).lower() in BLANK_TOKENS


def clean_text_value(value: Any) -> Any:
    text = norm_cell(value)
    if text.lower() in BLANK_TOKENS:
        return None
    return text


def map_dataframe(df: pd.DataFrame, func):
    mapper = getattr(df, "map", None)
    if mapper is not None:
        return mapper(func)
    return df.applymap(func)


def snake_case(text: str) -> str:
    text = norm_cell(text).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "column"


def display_header(text: str) -> str:
    text = norm_cell(text)
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return "Column"
    keep_upper = {
        "id",
        "sku",
        "arr",
        "mrr",
        "acv",
        "tcv",
        "cogs",
        "ebitda",
        "api",
        "sla",
        "po",
        "gl",
        "fy",
    }
    words = []
    for word in text.split(" "):
        stripped = re.sub(r"[^A-Za-z0-9]", "", word).lower()
        if stripped in keep_upper:
            words.append(stripped.upper())
        else:
            words.append(word[:1].upper() + word[1:].lower() if word else word)
    return " ".join(words)


def unique_headers(
    raw_headers: Iterable[Any], style: str
) -> tuple[list[str], list[dict[str, Any]]]:
    seen: dict[str, int] = defaultdict(int)
    headers: list[str] = []
    notes: list[dict[str, Any]] = []
    for idx, raw in enumerate(raw_headers, start=1):
        original = norm_cell(raw)
        if not original:
            original = f"column_{idx}"
            notes.append({"issue": "blank_header", "column_index": idx, "cleaned": original})
        base = snake_case(original) if style == "snake" else display_header(original)
        key = base.lower()
        seen[key] += 1
        cleaned = (
            base
            if seen[key] == 1
            else f"{base}_{seen[key]}"
            if style == "snake"
            else f"{base} {seen[key]}"
        )
        if seen[key] > 1:
            notes.append(
                {
                    "issue": "duplicate_header",
                    "column_index": idx,
                    "original": original,
                    "cleaned": cleaned,
                }
            )
        headers.append(cleaned)
    return headers, notes


def cell_has_letters(text: str) -> bool:
    return bool(re.search(r"[A-Za-z]", text))


def looks_numeric(text: str) -> bool:
    if not text:
        return False
    return bool(NUMERIC_RE.match(text.replace("%", "")))


def header_score(df: pd.DataFrame, row_idx: int) -> float:
    row = [norm_cell(v) for v in df.iloc[row_idx].tolist()]
    nonempty = [v for v in row if v]
    if len(nonempty) < 2:
        return -1000.0
    unique_ratio = len({v.lower() for v in nonempty}) / max(len(nonempty), 1)
    text_ratio = sum(cell_has_letters(v) for v in nonempty) / max(len(nonempty), 1)
    numeric_ratio = sum(looks_numeric(v) for v in nonempty) / max(len(nonempty), 1)
    next_ratio = 0.0
    if row_idx + 1 < len(df):
        next_vals = [norm_cell(v) for v in df.iloc[row_idx + 1].tolist()]
        next_ratio = sum(bool(v) for v in next_vals) / max(len(next_vals), 1)
    total_penalty = 2.0 if any(TOTAL_RE.search(v) for v in nonempty) else 0.0
    return (
        (len(nonempty) * 0.15)
        + (unique_ratio * 2.0)
        + (text_ratio * 3.0)
        + (next_ratio * 1.5)
        - (numeric_ratio * 2.0)
        - total_penalty
    )


def detect_header_row(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    scores = [(header_score(df, i), i) for i in range(min(len(df), 30))]
    scores.sort(reverse=True)
    return scores[0][1]


def read_input(path: Path, sheet: str | None = None) -> dict[str, pd.DataFrame]:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xls"}:
        data = pd.read_excel(
            path,
            sheet_name=sheet if sheet else None,
            header=None,
            dtype=object,
            keep_default_na=False,
        )
        if isinstance(data, pd.DataFrame):
            return {sheet or "sheet1": data}
        return {str(k): v for k, v in data.items()}
    if ext in {".csv", ".tsv", ".txt"}:
        sep = "\t" if ext == ".tsv" else None
        try:
            df = pd.read_csv(
                path,
                header=None,
                dtype=object,
                keep_default_na=False,
                sep=sep,
                engine="python",
                encoding="utf-8-sig",
            )
        except pd.errors.ParserError:
            df = pd.read_csv(
                path,
                header=None,
                dtype=object,
                keep_default_na=False,
                sep=sep,
                engine="python",
                encoding="utf-8-sig",
                on_bad_lines="warn",
            )
        return {"csv": df}
    raise ValueError(f"unsupported input type: {ext}")


def parse_number(value: Any) -> float | None:
    s = norm_cell(value)
    if not s or s.lower() in BLANK_TOKENS:
        return None
    s = CURRENCY_RE.sub("", s)
    s = s.replace(",", "").replace("%", "").strip()
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("() ")
    try:
        number = float(s)
        return -number if neg else number
    except ValueError:
        return None


def parse_percent(value: Any) -> float | None:
    s = norm_cell(value).lower()
    if not s or s in BLANK_TOKENS:
        return None
    if "bps" in s or "basis" in s:
        n = parse_number(s.replace("basis points", "").replace("bps", ""))
        return None if n is None else n / 10000.0
    n = parse_number(s)
    if n is None:
        return None
    if "%" in s:
        return n / 100.0
    # Leave ambiguous bare numbers as-is only if they already look like decimal percentages.
    return n if abs(n) <= 1 else None


def parse_bool(value: Any) -> bool | None:
    s = norm_cell(value).lower()
    if s in {"true", "yes", "y", "1"}:
        return True
    if s in {"false", "no", "n", "0"}:
        return False
    return None


def parse_dates(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def infer_column_type(header: str, series: pd.Series) -> dict[str, Any]:
    values = [norm_cell(v) for v in series.tolist()]
    nonblank = [v for v in values if v.lower() not in BLANK_TOKENS]
    sample_n = max(min(len(nonblank), 1000), 1)
    if not nonblank:
        return {"type": "empty", "confidence": 1.0, "format": "general"}

    header_id_like = bool(ID_HINT_RE.search(header))
    leading_zero = any(re.match(r"^0\d+", v) for v in nonblank[:1000])
    bool_ratio = sum(parse_bool(v) is not None for v in nonblank[:1000]) / sample_n
    pct_ratio = sum(parse_percent(v) is not None for v in nonblank[:1000]) / sample_n
    num_ratio = sum(parse_number(v) is not None for v in nonblank[:1000]) / sample_n
    currency_ratio = sum(bool(CURRENCY_RE.search(v)) for v in nonblank[:1000]) / sample_n
    date_candidates = [v for v in nonblank[:1000] if not looks_numeric(v)]
    if date_candidates:
        date_ratio = parse_dates(pd.Series(date_candidates)).notna().sum() / len(date_candidates)
    else:
        date_ratio = 0.0

    if header_id_like or leading_zero:
        return {"type": "text_identifier", "confidence": 0.90, "format": "@"}
    if bool_ratio >= 0.90:
        return {"type": "boolean", "confidence": bool_ratio, "format": "general"}
    if pct_ratio >= 0.80:
        return {"type": "percent", "confidence": pct_ratio, "format": "0.0%"}
    if currency_ratio >= 0.40 and num_ratio >= 0.80:
        return {
            "type": "currency_amount",
            "confidence": min(0.95, (currency_ratio + num_ratio) / 2),
            "format": "#,##0.00",
        }
    if num_ratio >= 0.88:
        decimals = any(
            (parse_number(v) is not None and abs(parse_number(v) - round(parse_number(v))) > 1e-9)
            for v in nonblank[:1000]
        )
        return {
            "type": "number",
            "confidence": num_ratio,
            "format": "#,##0.00" if decimals else "#,##0",
        }
    if date_ratio >= 0.85:
        return {"type": "date_or_datetime", "confidence": date_ratio, "format": "yyyy-mm-dd"}
    if 0.30 <= num_ratio < 0.88 or 0.30 <= date_ratio < 0.85:
        return {"type": "mixed", "confidence": max(num_ratio, date_ratio), "format": "general"}
    return {"type": "text", "confidence": 0.70, "format": "general"}


def convert_column(series: pd.Series, col_type: str) -> pd.Series:
    cleaned = series.map(clean_text_value)
    if col_type in {"text", "text_identifier", "mixed", "empty"}:
        return cleaned
    if col_type == "boolean":
        return cleaned.map(lambda v: parse_bool(v) if v is not None else None)
    if col_type == "percent":
        return cleaned.map(lambda v: parse_percent(v) if v is not None else None)
    if col_type in {"number", "currency_amount"}:
        return cleaned.map(lambda v: parse_number(v) if v is not None else None)
    if col_type == "date_or_datetime":
        return parse_dates(cleaned)
    return cleaned


def likely_subtotal_mask(df: pd.DataFrame) -> pd.Series:
    if df.empty:
        return pd.Series([], dtype=bool)

    def row_has_total(row: pd.Series) -> bool:
        vals = [norm_cell(v) for v in row.tolist()[:5]]
        return any(TOTAL_RE.search(v) for v in vals if v)

    return df.apply(row_has_total, axis=1)


def infer_domain(sheet_name: str, headers: list[str], override: str) -> str:
    if override and override != "auto":
        return override
    text = " ".join([sheet_name] + headers).lower()
    scores: dict[str, int] = {}
    for domain, hints in DOMAIN_REQUIRED_HINTS.items():
        scores[domain] = sum(1 for h in hints if re.search(r"\b" + re.escape(h) + r"\b", text))
    best, score = max(scores.items(), key=lambda x: x[1])
    return best if score else "general"


def required_field_checks(
    domain: str, clean_df: pd.DataFrame, sheet_name: str
) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    hints = DOMAIN_REQUIRED_HINTS.get(domain, [])
    lower_cols = {c.lower(): c for c in clean_df.columns}
    for hint in hints:
        matches = [actual for lower, actual in lower_cols.items() if hint in lower]
        for col in matches[:2]:
            missing = int(clean_df[col].isna().sum())
            if missing:
                checks.append(
                    {
                        "severity": "warning",
                        "issue_type": "missing_domain_field_values",
                        "sheet": sheet_name,
                        "field": col,
                        "affected_count": missing,
                        "description": f"{missing} rows are missing values in a domain-relevant field.",
                        "recommended_action": "review whether these rows are incomplete or need enrichment from the source system.",
                    }
                )
    return checks


def clean_sheet(
    sheet_name: str, raw_df: pd.DataFrame, args: argparse.Namespace
) -> tuple[pd.DataFrame, list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], str]:
    audit: list[dict[str, Any]] = []
    checks: list[dict[str, Any]] = []
    dictionary: list[dict[str, Any]] = []

    header_idx = args.header_row - 1 if args.header_row else detect_header_row(raw_df)
    raw_headers = raw_df.iloc[header_idx].tolist() if len(raw_df) else []
    headers, header_notes = unique_headers(raw_headers, args.header_style)
    data = raw_df.iloc[header_idx + 1 :].copy() if len(raw_df) else pd.DataFrame()
    if len(headers) == data.shape[1]:
        data.columns = headers

    audit.append(
        {
            "step": "header_detection",
            "action": f"used row {header_idx + 1} as header",
            "basis": "user-specified header row"
            if args.header_row
            else "automatic header inference",
            "affected_sheet": sheet_name,
            "affected_field": "all",
            "affected_rows_or_count": 1,
            "risk_level": "low" if args.header_row else "medium",
            "notes": "verify if the source has multi-row headers or report titles.",
        }
    )
    for note in header_notes:
        checks.append(
            {
                "severity": "warning",
                "issue_type": note["issue"],
                "sheet": sheet_name,
                "field": note.get("cleaned"),
                "affected_count": 1,
                "description": str(note),
                "recommended_action": "review header mapping in data_dictionary.",
            }
        )

    # Remove fully empty rows and columns from clean data.
    before_rows, before_cols = data.shape
    data = data.dropna(how="all")
    data = data.loc[:, [not all(is_blank(v) for v in data[col].tolist()) for col in data.columns]]
    # Convert blank tokens before another empty-row check.
    data = map_dataframe(data, clean_text_value)
    data = data.dropna(how="all")
    after_rows, after_cols = data.shape
    if before_rows != after_rows or before_cols != after_cols:
        audit.append(
            {
                "step": "structural_cleanup",
                "action": "removed fully empty rows/columns from clean output",
                "basis": "safe structural cleanup; raw_source remains preserved",
                "affected_sheet": sheet_name,
                "affected_field": "all",
                "affected_rows_or_count": f"rows {before_rows}->{after_rows}, columns {before_cols}->{after_cols}",
                "risk_level": "low",
                "notes": "only fully empty rows/columns were removed.",
            }
        )

    domain = infer_domain(sheet_name, list(data.columns), args.domain)
    subtotal_mask = likely_subtotal_mask(data)
    subtotal_count = int(subtotal_mask.sum())
    if subtotal_count:
        checks.append(
            {
                "severity": "warning",
                "issue_type": "possible_total_or_subtotal_rows",
                "sheet": sheet_name,
                "field": "row",
                "affected_count": subtotal_count,
                "description": "rows contain labels such as total or subtotal.",
                "recommended_action": "remove from detail data only if they are report subtotal rows, not real records.",
            }
        )
        if args.remove_subtotals:
            data = data.loc[~subtotal_mask].copy()
            audit.append(
                {
                    "step": "subtotal_handling",
                    "action": "removed possible total/subtotal rows",
                    "basis": "--remove-subtotals option",
                    "affected_sheet": sheet_name,
                    "affected_field": "row",
                    "affected_rows_or_count": subtotal_count,
                    "risk_level": "medium",
                    "notes": "verify these were not legitimate records.",
                }
            )

    # Infer and convert columns.
    for original_col in list(data.columns):
        info = infer_column_type(original_col, data[original_col])
        converted = convert_column(data[original_col], info["type"])
        data[original_col] = converted
        non_null = data[original_col].dropna()
        examples = [str(v) for v in non_null.head(5).tolist()]
        dictionary.append(
            {
                "source_sheet": sheet_name,
                "original_field": original_col,
                "clean_field": original_col,
                "inferred_type": info["type"],
                "excel_format": info["format"],
                "null_count": int(data[original_col].isna().sum()),
                "unique_count": int(non_null.nunique(dropna=True)) if len(non_null) else 0,
                "example_values": ", ".join(examples),
                "cleaning_notes": f"confidence={round(float(info['confidence']), 3)}",
                "business_notes": "",
            }
        )
        if info["type"] == "mixed":
            checks.append(
                {
                    "severity": "warning",
                    "issue_type": "mixed_type_column",
                    "sheet": sheet_name,
                    "field": original_col,
                    "affected_count": int(len(non_null)),
                    "description": "column contains mixed values that were left mostly as text.",
                    "recommended_action": "review source values and provide a conversion rule if this field should be numeric/date.",
                }
            )

    if args.dedupe == "exact" and not data.empty:
        before = len(data)
        data = data.drop_duplicates(keep="first")
        removed = before - len(data)
        if removed:
            audit.append(
                {
                    "step": "duplicate_handling",
                    "action": "removed exact duplicate rows",
                    "basis": "--dedupe exact; all cleaned values matched",
                    "affected_sheet": sheet_name,
                    "affected_field": "all",
                    "affected_rows_or_count": removed,
                    "risk_level": "low",
                    "notes": "potential non-exact duplicates are not merged by this script.",
                }
            )
            checks.append(
                {
                    "severity": "info",
                    "issue_type": "exact_duplicate_rows_removed",
                    "sheet": sheet_name,
                    "field": "all",
                    "affected_count": removed,
                    "description": "exact duplicates were removed from clean output.",
                    "recommended_action": "review if source row identity matters.",
                }
            )

    checks.extend(required_field_checks(domain, data, sheet_name))
    return data.reset_index(drop=True), dictionary, checks, audit, domain


def safe_excel_sheet_name(base: str, used: set[str]) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", base).strip() or "sheet"
    name = name[:31]
    original = name
    i = 2
    while name in used:
        suffix = f"_{i}"
        name = (original[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(name)
    return name


def autosize_and_table(path: Path, sheet_formats: dict[str, dict[str, str]]) -> None:
    load_workbook, Table, TableStyleInfo, get_column_letter = load_openpyxl_modules()
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= 1 and max_col >= 1:
            for cell in ws[1]:
                cell.style = "Headline 4"
            # Add a table where there is at least one data row and headers are populated.
            if max_row >= 2 and all(
                ws.cell(1, c).value not in (None, "") for c in range(1, max_col + 1)
            ):
                ref = f"A1:{get_column_letter(max_col)}{max_row}"
                table_name = re.sub(r"[^A-Za-z0-9_]", "_", f"tbl_{ws.title}")[:240]
                if not table_name or table_name[0].isdigit():
                    table_name = f"tbl_{table_name}"
                tab = Table(displayName=table_name, ref=ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                tab.tableStyleInfo = style
                try:
                    ws.add_table(tab)
                except ValueError:
                    pass
        formats = sheet_formats.get(ws.title, {})
        header_to_col = {str(ws.cell(1, c).value): c for c in range(1, max_col + 1)}
        for header, fmt in formats.items():
            col_idx = header_to_col.get(header)
            if col_idx:
                for row in range(2, max_row + 1):
                    ws.cell(row, col_idx).number_format = fmt
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 10
            for row in range(1, min(max_row, 200) + 1):
                value = ws.cell(row, col_idx).value
                if value is not None:
                    max_len = max(max_len, min(len(str(value)), 60))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 64)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Conservatively clean CSV/XLSX data into an audit-ready workbook."
    )
    parser.add_argument("input", help="input .xlsx, .xls, .csv, .tsv, or .txt file")
    parser.add_argument("--output", "-o", default="cleaned.xlsx", help="output .xlsx path")
    parser.add_argument("--sheet", help="optional Excel sheet name to clean")
    parser.add_argument(
        "--domain",
        default="auto",
        help="auto, finance, capiq_factset_export, cim_table, debt_schedule, process_tracker, qoe_schedule, investing, operations, sales, product, hr, procurement, or general",
    )
    parser.add_argument(
        "--dedupe",
        choices=["none", "exact"],
        default="exact",
        help="duplicate policy for clean output",
    )
    parser.add_argument("--header-row", type=int, help="1-based header row override")
    parser.add_argument(
        "--header-style",
        choices=["display", "snake"],
        default="display",
        help="cleaned header style",
    )
    parser.add_argument(
        "--remove-subtotals",
        action="store_true",
        help="remove rows that appear to be totals/subtotals",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    try:
        load_openpyxl_modules()
    except DependencyPreflightError as exc:
        parser.exit(2, f"{parser.prog}: error: {exc}\n")
    tables = read_input(input_path, args.sheet)

    all_dictionary: list[dict[str, Any]] = []
    all_checks: list[dict[str, Any]] = []
    all_audit: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []
    sheet_formats: dict[str, dict[str, str]] = {}
    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {
                    "item": "First read",
                    "detail": "Use this Cover tab first; cleaned data, raw source, QA, and audit tabs follow.",
                },
                {"item": "Input", "detail": str(input_path)},
                {"item": "Source tables", "detail": len(tables)},
                {
                    "item": "Cleaner posture",
                    "detail": "conservative first-pass cleaning; ambiguous business rules remain in quality_checks and assumptions_audit",
                },
                {
                    "item": "Support artifacts",
                    "detail": "No raw CSV/JSON support files are user-facing by default; manifest.json is an agent audit file.",
                },
            ]
        ).to_excel(writer, sheet_name="Cover", index=False)
        for source_name, raw_df in tables.items():
            clean_df, dictionary, checks, audit, domain = clean_sheet(source_name, raw_df, args)
            clean_sheet_name = safe_excel_sheet_name(
                "clean_data" if len(tables) == 1 else f"clean_{source_name}", used_sheet_names
            )
            raw_sheet_name = safe_excel_sheet_name(
                "raw_source" if len(tables) == 1 else f"raw_{source_name}", used_sheet_names
            )
            clean_df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
            raw_df.to_excel(writer, sheet_name=raw_sheet_name, index=False, header=False)
            fmt_map = {
                row["clean_field"]: row["excel_format"]
                for row in dictionary
                if row.get("excel_format") not in {None, "general"}
            }
            sheet_formats[clean_sheet_name] = fmt_map
            all_dictionary.extend(dictionary)
            all_checks.extend(checks)
            all_audit.extend(audit)
            summary_rows.append(
                {
                    "source_sheet": source_name,
                    "clean_sheet": clean_sheet_name,
                    "raw_sheet": raw_sheet_name,
                    "inferred_domain": domain,
                    "clean_rows": len(clean_df),
                    "clean_columns": len(clean_df.columns),
                    "quality_issue_count": len(checks),
                }
            )

        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="summary", index=False)
        pd.DataFrame(all_dictionary or [{"note": "no columns profiled"}]).to_excel(
            writer, sheet_name="data_dictionary", index=False
        )
        pd.DataFrame(
            all_checks
            or [
                {
                    "severity": "info",
                    "issue_type": "no_issues_logged",
                    "description": "no quality issues were logged by the deterministic cleaner",
                }
            ]
        ).to_excel(writer, sheet_name="quality_checks", index=False)
        all_audit.append(
            {
                "step": "workbook_creation",
                "action": "created cleaned workbook",
                "basis": "deterministic first-pass cleaning script",
                "affected_sheet": "all",
                "affected_field": "all",
                "affected_rows_or_count": len(tables),
                "risk_level": "low",
                "notes": f"generated_at={datetime.now(timezone.utc).isoformat()}; input={input_path}",
            }
        )
        pd.DataFrame(all_audit).to_excel(writer, sheet_name="assumptions_audit", index=False)

    autosize_and_table(output_path, sheet_formats)
    write_artifact_manifest(
        output_path.parent,
        "excel-data-cleaner",
        "workbook",
        output_path,
        extra={
            "inputs": {
                "input": str(input_path),
                "sheet": args.sheet or "",
                "domain": args.domain,
                "dedupe": args.dedupe,
            }
        },
    )
    print(f"wrote cleaned workbook to {output_path}")


if __name__ == "__main__":
    main()
