#!/usr/bin/env python3
"""
Static workbook inspection helper for the model-audit-tieout skill.

This script does not calculate Excel formulas. It inventories workbook structure,
formulas, visible/hidden sheets, hardcodes inside formulas, external links,
volatile functions, and simple formula-family inconsistencies. Its workbook output
is a mechanical screen, not a decision-readiness audit; apply financial-modeling
judgment and source tie-outs separately.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import warnings
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Iterable

PLUGIN_ROOT = Path(__file__).resolve().parents[3]
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from shared.artifacts import (  # noqa: E402
    artifact_item,
    dict_rows_to_sheet,
    support_dir,
    write_artifact_manifest,
    write_cover_first_workbook,
)

VOLATILE_FUNCTIONS = [
    "TODAY",
    "NOW",
    "RAND",
    "RANDBETWEEN",
    "OFFSET",
    "INDIRECT",
    "CELL",
    "INFO",
]

ERROR_LITERALS = ["#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"]
COMMON_CONSTANTS = {0, 1, -1, 2, -2, 10, 12, 52, 100, 365, 360, 1000, 1000000}

CELL_REF_RE = re.compile(r"(?<![A-Z0-9_])\$?[A-Z]{1,3}\$?\d+(?![A-Z0-9_])")
RANGE_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+\s*:\s*\$?[A-Z]{1,3}\$?\d+", re.IGNORECASE)
SHEET_REF_RE = re.compile(r"(?:'[^']+'|[A-Za-z0-9_ ]+)!\$?[A-Z]{1,3}\$?\d+", re.IGNORECASE)
EXTERNAL_REF_RE = re.compile(
    r"\[[^\]]+\]|\.xlsx|\.xlsm|\.xls|http[s]?://|\\\\|/[A-Za-z0-9_. -]+/", re.IGNORECASE
)
NUMERIC_LITERAL_RE = re.compile(r"(?<![A-Za-z0-9_\$])[-+]?\d+(?:\.\d+)?%?(?![A-Za-z0-9_])")
DATE_HINT_RE = re.compile(
    r"\b(as of|updated|last updated|date|period|source date|market data)\b", re.IGNORECASE
)


class MissingOptionalDependencyError(RuntimeError):
    """Raised when workbook inspection cannot run because an optional package is absent."""


def require_openpyxl_load_workbook() -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        install_command = "python3 -m pip install -r scripts/requirements.txt"
        requirements_path = Path(__file__).with_name("requirements.txt")
        raise MissingOptionalDependencyError(
            "\n".join(
                [
                    "ERROR: openpyxl is required for workbook inspection, but it is not installed.",
                    "Install the local script dependencies from the model-audit-tieout skill directory:",
                    f"  {install_command}",
                    f"Requirements file: {requirements_path}",
                    "Manual fallback: perform the SKILL.md manual model review workflow covering workbook structure, formulas, source tie-outs, assumptions, sensitivities, and decision-readiness.",
                    f"Import detail: {exc}",
                ]
            )
        ) from exc
    return load_workbook


def format_mechanical_screen_workbook(workbook_path: Path) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        workbook = require_openpyxl_load_workbook()(workbook_path)
    cover = workbook["Cover"]
    cover.merge_cells("A1:B1")
    cover["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    cover["A1"].fill = PatternFill("solid", fgColor="172B4D")
    cover["A1"].alignment = Alignment(vertical="center")
    cover.row_dimensions[1].height = 32
    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["B"].width = 96
    for row in range(2, 7):
        cover[f"A{row}"].font = Font(bold=True, color="334155")
        cover[f"A{row}"].alignment = Alignment(vertical="top")
        cover[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.row_dimensions[2].height = 34
    cover.row_dimensions[6].height = 34

    widths = {
        "Issues": [12, 22, 22, 18, 58, 58, 58],
        "Workbook_Map": [24, 14, 12, 12, 12, 14, 14, 18, 14, 16, 16, 18],
        "Formulas": [22, 14, 68, 68, 14, 20, 18, 20, 18, 22],
        "Sources": [22, 100],
    }
    for sheet_name, column_widths in widths.items():
        sheet = workbook[sheet_name]
        sheet.freeze_panes = "A2"
        for index, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="172B4D")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(workbook_path)


@dataclass
class WorkbookMapRow:
    sheet: str
    state: str
    max_row: int
    max_column: int
    used_cells: int
    formula_cells: int
    constant_cells: int
    blank_cells_in_used_range: int
    merged_ranges: int
    freeze_panes: str
    has_tables: bool
    has_auto_filter: bool


@dataclass
class FormulaRow:
    sheet: str
    cell: str
    formula: str
    normalized_formula: str
    formula_length: int
    cached_value: str
    has_external_ref: bool
    volatile_functions: str
    has_error_literal: bool
    hardcoded_numbers: str


@dataclass
class IssueRow:
    severity: str
    category: str
    sheet: str
    cell_or_range: str
    finding: str
    why_it_matters: str
    recommended_next_step: str


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def strip_strings(formula: str) -> str:
    # Replace double-quoted strings to avoid false positives in formulas.
    return re.sub(r'"(?:[^"]|"")*"', '"STR"', formula)


def normalize_formula(formula: str) -> str:
    s = strip_strings(formula.upper())
    s = RANGE_RE.sub("RANGE", s)
    s = SHEET_REF_RE.sub("SHEET!REF", s)
    s = CELL_REF_RE.sub("REF", s)
    s = re.sub(r"\s+", "", s)
    return s


def volatile_functions(formula: str) -> list[str]:
    found = []
    upper = formula.upper()
    for fn in VOLATILE_FUNCTIONS:
        if re.search(r"\b" + re.escape(fn) + r"\s*\(", upper):
            found.append(fn)
    return found


def extract_hardcoded_numbers(formula: str) -> list[str]:
    s = strip_strings(formula.upper())
    # Remove cell refs/ranges first to avoid row numbers appearing as numeric literals.
    s = RANGE_RE.sub(" RANGE ", s)
    s = SHEET_REF_RE.sub(" SHEET_REF ", s)
    s = CELL_REF_RE.sub(" REF ", s)
    raw = NUMERIC_LITERAL_RE.findall(s)
    values = []
    for item in raw:
        cleaned = item.rstrip("%")
        try:
            num = float(cleaned)
        except ValueError:
            continue
        if num.is_integer():
            comp = int(num)
        else:
            comp = num
        if comp in COMMON_CONSTANTS:
            continue
        # Keep percentages and non-common numbers; these are often assumptions.
        values.append(item)
    # Deduplicate while preserving order.
    seen = set()
    result = []
    for val in values:
        if val not in seen:
            seen.add(val)
            result.append(val)
    return result


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def add_issue(
    issues: list[IssueRow],
    severity: str,
    category: str,
    sheet: str,
    cell_or_range: str,
    finding: str,
    why: str,
    next_step: str,
) -> None:
    issues.append(
        IssueRow(
            severity=severity,
            category=category,
            sheet=sheet,
            cell_or_range=cell_or_range,
            finding=finding,
            why_it_matters=why,
            recommended_next_step=next_step,
        )
    )


def scan_workbook(
    workbook_path: Path,
) -> tuple[list[WorkbookMapRow], list[FormulaRow], list[IssueRow], dict[str, Any]]:
    load_workbook = require_openpyxl_load_workbook()
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        wb_formula = load_workbook(workbook_path, data_only=False, keep_links=True)
        try:
            wb_values = load_workbook(workbook_path, data_only=True, keep_links=True)
        except Exception:
            wb_values = None

    workbook_map: list[WorkbookMapRow] = []
    formulas: list[FormulaRow] = []
    issues: list[IssueRow] = []
    date_hints: list[dict[str, str]] = []
    formula_cells_by_sheet_row: dict[tuple[str, int], list[tuple[str, str, str]]] = defaultdict(
        list
    )

    external_links_count = len(getattr(wb_formula, "_external_links", []) or [])
    if external_links_count:
        add_issue(
            issues,
            "high",
            "formula_integrity",
            "workbook",
            "external links",
            f"workbook contains {external_links_count} external link object(s)",
            "external links can make audit results stale or unreproducible if source files are missing",
            "review linked files, refresh status, and tie material linked values to uploaded sources",
        )

    if getattr(wb_formula.calculation, "iterate", False):
        add_issue(
            issues,
            "high",
            "model_architecture",
            "workbook",
            "calculation settings",
            "iterative calculation appears enabled",
            "iterative settings can hide circularity in interest, debt, cash sweep, or plug logic",
            "document intentional circular references and test model behavior with iteration settings confirmed",
        )

    for ws in wb_formula.worksheets:
        ws_values = wb_values[ws.title] if wb_values and ws.title in wb_values.sheetnames else None
        used_cells = 0
        formula_cells = 0
        constant_cells = 0
        blank_cells = 0

        if ws.sheet_state != "visible":
            sev = "high" if ws.sheet_state == "veryHidden" else "medium"
            add_issue(
                issues,
                sev,
                "model_architecture",
                ws.title,
                "sheet visibility",
                f"sheet is {ws.sheet_state}",
                "hidden sheets can contain source data, assumptions, checks, or formulas not visible to reviewers",
                "inspect hidden sheet contents and disclose whether it feeds material outputs",
            )

        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                used_cells += 1
                coord = cell.coordinate
                if is_formula(val):
                    formula_cells += 1
                    cached = ""
                    if ws_values is not None:
                        try:
                            cached = stringify(ws_values[coord].value)
                        except Exception:
                            cached = ""
                    norm = normalize_formula(val)
                    vols = volatile_functions(val)
                    hardcodes = extract_hardcoded_numbers(val)
                    has_external = bool(EXTERNAL_REF_RE.search(val))
                    has_error = any(err in val.upper() for err in ERROR_LITERALS)
                    formulas.append(
                        FormulaRow(
                            sheet=ws.title,
                            cell=coord,
                            formula=val,
                            normalized_formula=norm,
                            formula_length=len(val),
                            cached_value=cached,
                            has_external_ref=has_external,
                            volatile_functions=";".join(vols),
                            has_error_literal=has_error,
                            hardcoded_numbers=";".join(hardcodes),
                        )
                    )
                    formula_cells_by_sheet_row[(ws.title, cell.row)].append((coord, val, norm))

                    if has_external:
                        add_issue(
                            issues,
                            "high",
                            "formula_integrity",
                            ws.title,
                            coord,
                            "formula contains an external reference",
                            "external references can produce stale or unreproducible values if the linked file is unavailable or outdated",
                            "tie the value to the linked source file or replace with documented source-tab data",
                        )
                    if vols:
                        add_issue(
                            issues,
                            "medium",
                            "formula_integrity",
                            ws.title,
                            coord,
                            f"formula uses volatile function(s): {', '.join(vols)}",
                            "volatile functions can change outputs on recalculation and complicate audit reproducibility",
                            "confirm volatility is intentional and does not drive material outputs without disclosure",
                        )
                    if hardcodes:
                        sev = "medium" if len(hardcodes) <= 2 else "high"
                        add_issue(
                            issues,
                            sev,
                            "formula_integrity",
                            ws.title,
                            coord,
                            f"formula contains review-worthy numeric literal(s): {', '.join(hardcodes[:8])}",
                            "hardcoded numbers inside formulas may hide assumptions, rates, multiples, or adjustments",
                            "move material assumptions to an input/source tab or document why the constant is appropriate",
                        )
                    if has_error:
                        add_issue(
                            issues,
                            "high",
                            "formula_integrity",
                            ws.title,
                            coord,
                            "formula contains an error literal",
                            "error literals may indicate broken references or intentionally suppressed errors",
                            "inspect precedent cells and resolve or document the error treatment",
                        )
                else:
                    constant_cells += 1
                    if isinstance(val, str) and DATE_HINT_RE.search(val):
                        # Capture local context around date/source labels for manual review.
                        date_hints.append({"sheet": ws.title, "cell": coord, "text": val[:200]})
                # blanks in used range are computed below

        total_range_cells = ws.max_row * ws.max_column if ws.max_row and ws.max_column else 0
        blank_cells = max(total_range_cells - used_cells, 0)
        workbook_map.append(
            WorkbookMapRow(
                sheet=ws.title,
                state=ws.sheet_state,
                max_row=ws.max_row,
                max_column=ws.max_column,
                used_cells=used_cells,
                formula_cells=formula_cells,
                constant_cells=constant_cells,
                blank_cells_in_used_range=blank_cells,
                merged_ranges=len(list(ws.merged_cells.ranges)),
                freeze_panes=stringify(ws.freeze_panes),
                has_tables=bool(getattr(ws, "tables", None)),
                has_auto_filter=bool(ws.auto_filter and ws.auto_filter.ref),
            )
        )

        if len(list(ws.merged_cells.ranges)) > 0:
            add_issue(
                issues,
                "low",
                "formatting",
                ws.title,
                "merged ranges",
                f"sheet contains {len(list(ws.merged_cells.ranges))} merged range(s)",
                "merged cells can complicate formula review, data extraction, and tie-out automation",
                "review whether merged ranges are cosmetic or interfere with source data / calculations",
            )

    # Identify simple formula-family inconsistencies by row.
    for (sheet, rownum), row_formulas in formula_cells_by_sheet_row.items():
        if len(row_formulas) < 4:
            continue
        norms = [item[2] for item in row_formulas]
        counts = Counter(norms)
        dominant_norm, dominant_count = counts.most_common(1)[0]
        # Flag minority formulas when a clear dominant pattern exists.
        if dominant_count >= max(3, int(len(row_formulas) * 0.65)) and len(counts) > 1:
            minority = [
                (coord, formula) for coord, formula, norm in row_formulas if norm != dominant_norm
            ]
            if 0 < len(minority) <= 5:
                cells = ", ".join(coord for coord, _ in minority)
                add_issue(
                    issues,
                    "medium",
                    "formula_integrity",
                    sheet,
                    f"row {rownum}: {cells}",
                    "row contains formula(s) that differ from the dominant copied formula pattern",
                    "unexpected formula-family breaks can indicate overwritten formulas, wrong period references, or hidden assumptions",
                    "inspect the minority cells and confirm whether the break is intentional",
                )

    # Flag sheets that are mostly constants but named like assumptions/source checks.
    for row in workbook_map:
        name = row.sheet.lower()
        if (
            row.used_cells
            and row.formula_cells == 0
            and any(token in name for token in ["source", "input", "assumption", "data"])
        ):
            add_issue(
                issues,
                "info",
                "source_tieout",
                row.sheet,
                "entire sheet",
                "sheet appears to be a source/input tab with no formulas",
                "source/input tabs should be tied to underlying documents and as-of dates",
                "verify source labels, source dates, and tie-outs for material values on this tab",
            )

    summary = {
        "workbook": str(workbook_path.name),
        "sheets": len(wb_formula.worksheets),
        "hidden_sheets": sum(1 for r in workbook_map if r.state != "visible"),
        "formula_cells": len(formulas),
        "external_link_objects": external_links_count,
        "issues_by_severity": dict(Counter(issue.severity for issue in issues)),
        "issues_by_category": dict(Counter(issue.category for issue in issues)),
        "date_source_hints": date_hints[:100],
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "notes": [
            "static inspection only; this script does not calculate formulas or prove business logic is correct",
            "use outputs as a triage starting point and supplement with manual model review",
        ],
    }
    return workbook_map, formulas, issues, summary


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Static audit of an Excel workbook for model-audit-tieout."
    )
    parser.add_argument("workbook", help="Path to .xlsx or .xlsm workbook")
    parser.add_argument(
        "--out-dir", default="audit_output", help="Directory where audit outputs will be written"
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        print(f"ERROR: workbook not found: {workbook_path}", file=sys.stderr)
        return 1
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        print(
            "ERROR: only .xlsx and .xlsm files are supported by this static inspector.",
            file=sys.stderr,
        )
        return 1

    out_dir = Path(args.out_dir).expanduser().resolve()

    try:
        workbook_map, formulas, issues, summary = scan_workbook(workbook_path)
    except MissingOptionalDependencyError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    out_dir.mkdir(parents=True, exist_ok=True)
    csv_dir = support_dir(out_dir)

    workbook_map_rows = [asdict(row) for row in workbook_map]
    formula_rows = [asdict(row) for row in formulas]
    issue_rows = [asdict(row) for row in issues]

    write_csv(
        csv_dir / "workbook_map.csv",
        workbook_map_rows,
        list(asdict(workbook_map[0]).keys()) if workbook_map else [],
    )
    if formulas:
        write_csv(csv_dir / "formulas.csv", formula_rows, list(asdict(formulas[0]).keys()))
    else:
        write_csv(
            csv_dir / "formulas.csv",
            [],
            [
                "sheet",
                "cell",
                "formula",
                "normalized_formula",
                "formula_length",
                "cached_value",
                "has_external_ref",
                "volatile_functions",
                "has_error_literal",
                "hardcoded_numbers",
            ],
        )
    if issues:
        write_csv(csv_dir / "issues.csv", issue_rows, list(asdict(issues[0]).keys()))
    else:
        write_csv(
            csv_dir / "issues.csv",
            [],
            [
                "severity",
                "category",
                "sheet",
                "cell_or_range",
                "finding",
                "why_it_matters",
                "recommended_next_step",
            ],
        )

    summary_path = csv_dir / "audit_summary.json"
    with summary_path.open("w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    audit_workbook = out_dir / "model_audit_screen.xlsx"
    write_cover_first_workbook(
        audit_workbook,
        [
            ["Model Audit Mechanical Screen"],
            ["Audited workbook", str(workbook_path)],
            ["Sheets", summary.get("sheets", 0)],
            ["Formula cells", summary.get("formula_cells", 0)],
            ["Issues", len(issues)],
            [
                "Reliance boundary",
                "Mechanical screen only; complete source tie-outs and judgmental output-path review before reliance.",
            ],
        ],
        {
            "Issues": dict_rows_to_sheet(
                issue_rows,
                list(asdict(issues[0]).keys())
                if issues
                else [
                    "severity",
                    "category",
                    "sheet",
                    "cell_or_range",
                    "finding",
                    "why_it_matters",
                    "recommended_next_step",
                ],
            ),
            "Workbook_Map": dict_rows_to_sheet(workbook_map_rows),
            "Formulas": dict_rows_to_sheet(formula_rows),
            "Sources": [
                ["source", "path"],
                ["audited_workbook", str(workbook_path)],
                ["support_folder", str(csv_dir)],
            ],
        },
    )
    format_mechanical_screen_workbook(audit_workbook)

    write_artifact_manifest(
        out_dir,
        "model-audit-tieout",
        "workbook",
        audit_workbook,
        support_artifacts=[
            artifact_item(
                csv_dir / "workbook_map.csv",
                "support_artifact",
                "csv",
                "Raw workbook map inventory.",
                False,
                True,
                "CSV is raw audit support for filtering/import.",
            ),
            artifact_item(
                csv_dir / "formulas.csv",
                "support_artifact",
                "csv",
                "Raw formula inventory.",
                False,
                True,
                "CSV is raw audit support for filtering/import.",
            ),
            artifact_item(
                csv_dir / "issues.csv",
                "support_artifact",
                "csv",
                "Raw issue table.",
                False,
                True,
                "CSV backs the mechanical-screen workbook and later judgmental review.",
            ),
            artifact_item(
                summary_path,
                "support_artifact",
                "json",
                "Raw audit summary JSON.",
                False,
                True,
                "JSON is machine-readable audit support.",
            ),
        ],
        blocked_or_partial_status={
            "status": "partial",
            "reason": "This deterministic workbook is a mechanical screen only; decision-readiness requires source tie-outs and judgmental review.",
            "missing_inputs": ["Source tie-outs and material output-path review"],
        },
    )

    print(f"Wrote audit outputs to: {out_dir}")
    print(
        f"Sheets: {summary.get('sheets', 0)} | Formula cells: {summary.get('formula_cells', 0)} | Issues: {len(issues)}"
    )
    print(f"Mechanical screen workbook: {audit_workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
